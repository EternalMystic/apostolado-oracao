"""Cria data/apostolado.xlsx com dados iniciais do seed."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from dados_membros import (
    CONFIG_PADRAO,
    INCONSISTENCIAS_SEED,
    ITENS_ENTREGA,
    MEMBROS_SEED,
    MEMORIAL,
    ORDEM_BAIRROS,
)
from data_manager import (
    COL_AGENDA,
    COL_CONFIG,
    COL_CONSAGRACOES,
    COL_ENTREGAS,
    COL_INCONSISTENCIAS,
    COL_INTENCOES,
    COL_MEMBROS,
    COL_MEMORIAL,
    COL_VISITAS,
    DATA_DIR,
    EXCEL_PATH,
    ROOT,
    SHEET_AGENDA,
    SHEET_CONFIG,
    SHEET_CONSAGRACOES,
    SHEET_ENTREGAS,
    SHEET_INCONSISTENCIAS,
    SHEET_INTENCOES,
    SHEET_MEMBROS,
    SHEET_MEMORIAL,
    SHEET_VISITAS,
)

COR_ROXO = "6A1B9A"


def _header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=COR_ROXO)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _df_membros() -> pd.DataFrame:
    return pd.DataFrame(
        [dict(zip(COL_MEMBROS, m)) for m in MEMBROS_SEED], columns=COL_MEMBROS
    )


def _df_inconsistencias() -> pd.DataFrame:
    return pd.DataFrame(
        [dict(zip(COL_INCONSISTENCIAS, t)) for t in INCONSISTENCIAS_SEED],
        columns=COL_INCONSISTENCIAS,
    )


def _df_config() -> pd.DataFrame:
    extra = {
        "itens_entrega": "|".join(ITENS_ENTREGA),
        "ordem_bairros": "|".join(ORDEM_BAIRROS),
    }
    merged = {**CONFIG_PADRAO, **extra}
    return pd.DataFrame(
        [{"chave": k, "valor": str(v)} for k, v in merged.items()],
        columns=COL_CONFIG,
    )


def _df_memorial() -> pd.DataFrame:
    rows = []
    for nome, nasc, falec, obs in MEMORIAL:
        rows.append(
            {
                "nome": nome,
                "nasc": nasc,
                "falecimento": falec,
                "observacao": obs,
            }
        )
    return pd.DataFrame(rows, columns=COL_MEMORIAL)


def _df_consagracoes() -> pd.DataFrame:
    rows = []
    rid = 1
    for m in MEMBROS_SEED:
        if str(m[11]).strip().lower() == "sim":
            rows.append(
                {
                    "id": rid,
                    "membro_id": m[0],
                    "membro_nome": m[2],
                    "data_consagracao": "",
                    "local": "Paróquia São Jorge",
                    "observacoes": m[12],
                }
            )
            rid += 1
    return pd.DataFrame(rows, columns=COL_CONSAGRACOES)


def _df_rota_entregas() -> pd.DataFrame:
    """Linhas iniciais da rota para membros ativos com endereço."""
    rows = []
    eid = 1
    ativos = [m for m in MEMBROS_SEED if m[10] in ("Ativo", "Ativo (presumido)")]
    for m in sorted(ativos, key=lambda x: (x[7] or "", x[2])):
        rows.append(
            {
                "id": eid,
                "membro_id": m[0],
                "membro_nome": m[2],
                "item": ITENS_ENTREGA[0],
                "data_entrega": "",
                "entregue": "N",
                "observacoes": m[7] or "",
            }
        )
        eid += 1
    return pd.DataFrame(rows, columns=COL_ENTREGAS)


def _sheet_from_df(ws, df: pd.DataFrame) -> None:
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws[1]:
        _header_style(cell)


def criar_workbook_inicial(dest: Path | None = None) -> Path:
    dest = dest or EXCEL_PATH
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (ROOT / "backups").mkdir(exist_ok=True)

    sheets = {
        SHEET_MEMBROS: _df_membros(),
        SHEET_INCONSISTENCIAS: _df_inconsistencias(),
        SHEET_ENTREGAS: _df_rota_entregas(),
        SHEET_VISITAS: pd.DataFrame(columns=COL_VISITAS),
        SHEET_CONSAGRACOES: _df_consagracoes(),
        SHEET_INTENCOES: pd.DataFrame(columns=COL_INTENCOES),
        SHEET_AGENDA: pd.DataFrame(columns=COL_AGENDA),
        SHEET_CONFIG: _df_config(),
        SHEET_MEMORIAL: _df_memorial(),
    }

    wb = Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        ws = wb.create_sheet(name)
        _sheet_from_df(ws, df)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    wb.save(dest)
    wb.close()
    return dest


def main():
    path = criar_workbook_inicial()
    print(f"Excel criado: {path}")
    print(f"Membros: {len(MEMBROS_SEED)} · Abas: 8")


if __name__ == "__main__":
    main()
